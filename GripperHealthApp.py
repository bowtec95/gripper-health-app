import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
import csv
import io

st.set_page_config(layout="wide")

st.image("mujin logo.png", width=250)

st.title("Gripper Health App")

# --- Gripper Selection ---
gripper_type = st.selectbox(
    "Select Gripper Type",
    ["(Select)", "63 Channel Gripper", "Mega Gripper"]
)

gripper_version = st.selectbox(
    "Select Gripper Version",
    ["(Select)", "V1", "V2", "V3"]
)

if gripper_type == "(Select)" or gripper_version == "(Select)":
    st.warning("Please select both Gripper Type and Version")
    st.stop()

st.write(f"Selected: {gripper_version} {gripper_type}")

# --- Upload File ---
uploaded_file = st.file_uploader(
    "Upload spreadsheet file",
    type=["xlsx", "csv"]
)

# --- Highlight Modules ---
highlight_text = st.text_input(
    "Highlight specific modules",
    placeholder="Example: 23,30,19,18"
)

highlight_modules = []

if highlight_text:
    for item in highlight_text.split(","):
        try:
            highlight_modules.append(int(item.strip()))
        except:
            pass

# --- Pressure Targets ---
if gripper_type == "Mega Gripper":

    TARGET_LOW = -500
    TARGET_HIGH = -300
    GRAPH_MIN = -700

else:

    TARGET_LOW = -40000
    TARGET_HIGH = -25000
    GRAPH_MIN = -70000

# --- Module Color Logic ---
def get_module_color(samples):

    core_samples = samples[1:7]

    min_target = min(TARGET_LOW, TARGET_HIGH)
    max_target = max(TARGET_LOW, TARGET_HIGH)

    for index, value in enumerate(core_samples):

        try:
            value = float(value)
        except:
            continue

        # In acceptable range
        if min_target <= value <= max_target:

            # Fast response
            if index <= 1:
                return "green"

            # Delayed response
            else:
                return "yellow"

    return "red"

# --- Status Text ---
def get_status_text(status):

    if status == "green":
        return "Reached acceptable pressure and maintained"

    elif status == "yellow":
        return "Delayed time in reaching acceptable pressure"

    elif status == "red":
        return "Failed to reach acceptable pressure"

    return "No data"

module_colors = {}
module_samples = {}

# --- Analyze Spreadsheet ---
if uploaded_file:

    file_name = uploaded_file.name.lower()

    # --- CSV ---
    if file_name.endswith(".csv"):

        text = uploaded_file.getvalue().decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)

        max_col = len(rows[0])

        for col in range(1, max_col + 1):

            try:
                module_name = rows[0][col - 1]
                module_number = int(float(module_name))
            except:
                continue

            module_start = 2 + ((module_number - 1) * 6)

            start_row = module_start - 1
            end_row = module_start + 7

            samples = []

            for row in range(start_row, end_row):

                try:
                    value = rows[row - 1][col - 1]
                except:
                    value = None

                if value not in [None, ""]:
                    samples.append(float(value))

            module_samples[module_number] = samples
            module_colors[module_number] = get_module_color(samples)

    # --- XLSX ---
    else:

        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active

        max_col = ws.max_column

        for col in range(1, max_col + 1):

            module_name = ws.cell(row=1, column=col).value

            try:
                module_number = int(float(module_name))
            except:
                continue

            module_start = 2 + ((module_number - 1) * 6)

            start_row = module_start - 1
            end_row = module_start + 7

            samples = []

            for row in range(start_row, end_row):

                value = ws.cell(row=row, column=col).value

                if value is not None:
                    samples.append(float(value))

            module_samples[module_number] = samples
            module_colors[module_number] = get_module_color(samples)

    st.success("File uploaded and analyzed successfully")

# --- Gripper Layout ---
if gripper_type in ["63 Channel Gripper", "Mega Gripper"]:

    st.subheader(f"{gripper_version} {gripper_type} Layout")

    if gripper_type == "63 Channel Gripper":

        rows = 7
        cols = 9
        box_size = 40
        grid_height = 450

    else:

        rows = 5
        cols = 22
        box_size = 40
        grid_height = 360

    html = f"""
    <div style="
        display: grid;
        grid-template-columns: repeat({cols}, {box_size}px);
        gap: 6px;
        padding: 12px;
        border: 3px solid black;
        width: max-content;
        background-color: #f5f5f5;
    ">
    """

    for row in range(rows):
        for col in range(cols):

            # Bottom-right = 1
            # Top-left = highest number
            num = (rows - row) + ((cols - 1 - col) * rows)

            color = module_colors.get(num, "white")

            # Highlight border
            if num in highlight_modules:
                inner_shadow = "inset 0 0 0 4px hotpink"
            else:
                inner_shadow = "none"

            html += f"""
            <div style="
                width: {box_size}px;
                height: {box_size}px;
                border: 1px solid black;
                box-shadow: {inner_shadow};
                background-color: {color};
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 9px;
                font-weight: bold;
            ">
            {num}
            </div>
            """

    html += "</div>"

    components.html(html, height=grid_height)

    st.caption("Orientation: Front face of gripper")

# --- Module Graph Section ---
if uploaded_file:

    max_module = 110 if gripper_type == "Mega Gripper" else 63

    selected_module = st.selectbox(
        "Select module to graph",
        list(range(1, max_module + 1))
    )

    samples = module_samples.get(selected_module, [])

    if len(samples) >= 7:

        graph_samples = [0] + samples[1:7] + [0]

        st.subheader(f"Module {selected_module} Pressure Readings")

        status = module_colors.get(selected_module, "white")

        status_text = get_status_text(status)

        st.write(f"Status: {status_text}")

        graph_width = 700
        graph_height = 420

        left_margin = 70
        top_margin = 30
        bottom_y = 360
        chart_height = 330
        point_spacing = 75

        def get_y(value):

            return bottom_y - (
                (0 - value) /
                (0 - GRAPH_MIN) * chart_height
            )

        points = []

        for i, value in enumerate(graph_samples):

            x = left_margin + (i * point_spacing)
            y = get_y(value)

            points.append(f"{x},{y}")

        polyline_points = " ".join(points)

        target_top_y = get_y(TARGET_LOW)
        target_bottom_y = get_y(TARGET_HIGH)

        target_height = (
            target_bottom_y - target_top_y
        )

        graph_html = f"""
        <svg width="{graph_width}"
             height="{graph_height}"
             style="
                border:1px solid black;
                background:#f9f9f9;
             ">

            <!-- Goal Window -->
            <rect x="{left_margin}"
                  y="{target_top_y}"
                  width="560"
                  height="{target_height}"
                  fill="lightgreen"
                  opacity="0.35" />

            <!-- Axes -->
            <line x1="{left_margin}"
                  y1="{bottom_y}"
                  x2="620"
                  y2="{bottom_y}"
                  stroke="black" />

            <line x1="{left_margin}"
                  y1="{top_margin}"
                  x2="{left_margin}"
                  y2="{bottom_y}"
                  stroke="black" />

            <!-- Data Line -->
            <polyline
                points="{polyline_points}"
                fill="none"
                stroke="blue"
                stroke-width="3"
            />
        """

        for i, value in enumerate(graph_samples):

            x = left_margin + (i * point_spacing)
            y = get_y(value)

            graph_html += f"""
            <circle cx="{x}"
                    cy="{y}"
                    r="4"
                    fill="red" />

            <text x="{x - 20}"
                  y="{y - 10}"
                  font-size="10">
                  {int(value)}
            </text>
            """

        x_labels = [
            "Start",
            "1",
            "2",
            "3",
            "4",
            "5",
            "6",
            "End"
        ]

        for i, label in enumerate(x_labels):

            x = left_margin + (i * point_spacing)

            graph_html += f"""
            <text x="{x - 10}"
                  y="390"
                  font-size="12">
                  {label}
            </text>
            """

        graph_html += f"""

            <!-- Y Labels -->
            <text x="25"
                  y="{bottom_y}"
                  font-size="12">
                  0
            </text>

            <text x="5"
                  y="{top_margin + 5}"
                  font-size="12">
                  {GRAPH_MIN}
            </text>

            <text x="5"
                  y="{get_y(TARGET_HIGH)}"
                  font-size="12">
                  {TARGET_HIGH}
            </text>

            <text x="5"
                  y="{get_y(TARGET_LOW)}"
                  font-size="12">
                  {TARGET_LOW}
            </text>

        </svg>
        """

        components.html(graph_html, height=440)

# --- PDF Report ---
def create_pdf_report():

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(
        Paragraph("Gripper Health Report", styles["Title"])
    )

    story.append(
        Paragraph(
            f"Gripper: {gripper_version} {gripper_type}",
            styles["Normal"]
        )
    )

    story.append(
        Paragraph(
            f"Target Range: {TARGET_LOW} to {TARGET_HIGH}",
            styles["Normal"]
        )
    )

    story.append(
        Paragraph(
            "Orientation: Front face of gripper",
            styles["Normal"]
        )
    )

    story.append(Spacer(1, 12))

    story.append(
        Paragraph("Gripper Layout", styles["Heading2"])
    )

    if gripper_type == "63 Channel Gripper":
        rows = 7
        cols = 9
    else:
        rows = 5
        cols = 22

    layout_data = []

    for row in range(rows):

        row_data = []

        for col in range(cols):

            num = (rows - row) + ((cols - 1 - col) * rows)

            row_data.append(str(num))

        layout_data.append(row_data)

    table = Table(layout_data)

    table_style = [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
    ]

    for row in range(rows):
        for col in range(cols):

            num = (rows - row) + ((cols - 1 - col) * rows)

            status = module_colors.get(num, "white")

            if status == "green":
                bg_color = colors.lightgreen
            elif status == "yellow":
                bg_color = colors.yellow
            elif status == "red":
                bg_color = colors.red
            else:
                bg_color = colors.white

            table_style.append(
                ("BACKGROUND", (col, row), (col, row), bg_color)
            )

            if num in highlight_modules:
                table_style.append(
                    ("BOX", (col, row), (col, row), 2, colors.magenta)
                )

    table.setStyle(TableStyle(table_style))

    story.append(table)

    story.append(Spacer(1, 16))

    issue_rows = [["Module", "Status", "Reason", "Samples"]]

    max_module = 110 if gripper_type == "Mega Gripper" else 63

    for module in range(1, max_module + 1):

        status = module_colors.get(module, "white")

        if status in ["yellow", "red"]:

            samples = module_samples.get(module, [])

            issue_rows.append([
                str(module),
                status.upper(),
                get_status_text(status),
                str(samples[1:7])
            ])

    story.append(
        Paragraph(
            "Modules Requiring Attention",
            styles["Heading2"]
        )
    )

    if len(issue_rows) == 1:

        story.append(
            Paragraph(
                "No delayed or failed modules found.",
                styles["Normal"]
            )
        )

    else:

        issue_table = Table(
            issue_rows,
            colWidths=[60, 80, 240, 300]
        )

        issue_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))

        story.append(issue_table)

    story.append(Spacer(1, 16))

    story.append(
        Paragraph(
            "Manually Flagged Modules for Inspection",
            styles["Heading2"]
        )
    )

    if highlight_modules:

        story.append(
            Paragraph(
                ", ".join(str(x) for x in highlight_modules),
                styles["Normal"]
            )
        )

    else:

        story.append(
            Paragraph("None", styles["Normal"])
        )

    doc.build(story)

    buffer.seek(0)

    return buffer

# --- PDF Export Button ---
if uploaded_file:

    st.download_button(
        label="Export PDF Report",
        data=create_pdf_report(),
        file_name="gripper_health_report.pdf",
        mime="application/pdf"
    )
